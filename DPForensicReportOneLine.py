#Convert DefensePro forensic with details to one line per entry
#Created and maintained by Steve Harris - Steven.Harris@radware.com
#Date Created: 14 March 2024
#Last Updated: 21 June 2024
#version 1.2
import os
import csv
import io
import re
import time
try:
    from benchmark import LoopTimer
except ImportError:
    class LoopTimer:
        def __init__(self, *args, **kwargs): pass
        def reset(self): pass
        def lap(self, *args, **kwargs): return 0
        def total(self, *args, **kwargs): return 0
BENCHMARK_ENABLED = False

from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl is not installed. Please install it by running: pip install openpyxl")
    exit()

input_path = "./input/"
output_path = "./output/"
replaceExistingFile = False
maxWidth=64
ColorAlternateRows = True
fillColor = 'F0F0F0F0'
hideEmptyColumns = True
DataHeaders = ["S.No","Start Time","End Time","Device IP Address","Threat Category","Attack Name","Policy Name","Action","Attack ID","Source IP Address","Source Port","Destination IP Address","Destination Port","Direction","Protocol","Radware ID","Duration","Total Packets","Total Packets Dropped","Packet Type","Total Mbits","Total Mbits Dropped","Max pps","Max bps","Max Attack Rate in Kb","Physical Port","Risk","VLAN Tag","Footprint","Device Name","Device Type","Workflow Rule Process","Activation Id","Protected Object"]

def rowSearch(rowStr,entry):
    try:
        return re.search(rf'(\n{re.escape(rowStr)})(.+)',entry,re.MULTILINE).group(2)
    except:
        #rowStr not found.
        return('')

class clsEntry:
    def __init__(self,rawEntry,startLine=1):
        #Grab header columns from the second row of the entry.
        self.data = []
        for i in range(len(DataHeaders)):
            self.data.append('')
        lines = rawEntry.splitlines()
        headerCount = 0
        for line in lines:
            if line.startswith("S.No,"):
                headerCount += 1
        if headerCount > 1:
            print(f"    Warning: Err1:Corrupt entry at line {startLine}. Multiple headers in same entry.")
            self.data[0] = f"Err1:{startLine}"
            #self.defaultHeader[0] = "Multiple in same entry:"
        elif headerCount == 0:
            print(f"    Warning: Err2:Corrupt entry at line {startLine}. Header line is missing.")
            self.data[0] = f"Err2:{startLine}"
            #self.defaultHeader[0] = "Header line missing"
        if len(lines) < 2:
            print(f"    Err3:Corrupt entry detected at line {startLine}. <2 lines in entry.")
            self.error = f"Err3:{startLine}"
            return
        
        matches = re.findall(r'(^S.No,.*\n)(.*$)',rawEntry, re.MULTILINE)
        for match in matches:
            reader = csv.reader(io.StringIO(match[0]))
            #headers = next(reader)
            headers = [h for h in next(reader) if h.strip() != '']
            reader = csv.reader(io.StringIO(match[1]))
            data = next(reader)
            if len(headers) > len(data): #We can't have more headers than entries
                if data[0] in ['SAMPLE DETAILS:', 'State']:
                    print(f"    Warning: Corrupt entry at line {startLine}. Missing or out of place data row.")
                elif data[0] in ['S.No']:
                    pass#second header. We've already notified the user of the issue. 
                else:
                    print(f"    Warning: Corrupt entry at line {startLine}. More headers than data:")
                    print(f'      Data: {data}\n      Headers: {headers}')
                self.data[0] = '\n    '.join([f"Err4:{startLine}",self.data[0]]).strip() 
            if len(headers) > 11:
                self.very_old_date_format = False
            try:
                if not data[0].isnumeric():
                    if not data[0].startswith('Err'):
                        pass
                        #self.defaultHeader[0] = f'Err5:{startLine}'
                        #raise ValueError("Err5:Bad data")
                else:
                    for i, header in enumerate(headers):
                        if data[i] != '':
                            index = DataHeaders.index(header)
                            self.data[index] = '\n    '.join([self.data[index],data[i]]).strip()
            except Exception as err:
                print(f"  Bad Data detected when parsing entry at line {startLine}: Headers: {headers} Data: {data}")
                print(f"      Details: ", err)
        #Parse specific rows
        self.footprint = rowSearch("Footprint,",rawEntry).strip('"')
        self.state = rowSearch("State,",rawEntry)        
        self.sIP = rowSearch("Source IP,",rawEntry)
        #If we matched the header row of the SAMPLE DETAILS section, the entry does not contain source ips.
        if self.sIP.startswith(" Source Port,"):
            self.sIP = ''
        self.sPort = rowSearch("Source Port,",rawEntry)
        self.dIP = rowSearch("Destination IP,",rawEntry)
        self.dPort = rowSearch("Destination Port,",rawEntry)

        #Parse Sample Data. DOTALL is so '.' matches newline
        sampleMatch = re.search(r'(^SAMPLE DETAILS:(?:,*)\n)(.*)',rawEntry,re.MULTILINE | re.DOTALL)
        self.samples=[]
        if sampleMatch:
            sampleLines = sampleMatch.group(2).splitlines()

            #Iterate through sampleLines starting with line 2. Line 1 is a header.
            for sampleLine in sampleLines[1:]:
                reader = csv.reader(io.StringIO(sampleLine))
                self.samples.append(next(reader))
        #else:
        #    print(f"SAMPLE DETAILS not found for id: {self.defaultHeader[8]}")




def processData(rawData):
    timer = LoopTimer(BENCHMARK_ENABLED)
    timer.reset()
    
    #Identify date format:
    def detect_date_format(raw_data):
        # Regex for dates like 13.09.2024 00:09:35 or 09.13.2024 12:00:00
        date_regex = re.compile(r'(\d{1,2})\.(\d{1,2})\.(\d{4})\s+(\d{2}):(\d{2}):(\d{2})')

        for match in date_regex.finditer(raw_data):
            d1, d2, year, h, m, s = match.groups()
            d1 = int(d1)
            d2 = int(d2)

            # Check for unambiguous day/month
            if d1 > 12 and d2 <= 12:
                return "%d.%m.%Y %H:%M:%S"
            elif d2 > 12 and d1 <= 12:
                return "%m.%d.%Y %H:%M:%S"
            elif d1 > 12 and d2 > 12:
                # Technically invalid date, skip
                continue

        # Default fallback (common case)
        return "%m.%d.%Y %H:%M:%S"
    dp_one_line_date_format = detect_date_format(rawData)
    timer.lap("Date format detected")

    #Initialize the workbook and sheet.
    wb = openpyxl.Workbook()
    sheet = wb.active

    #Create the header row. Include all our DataHeaders and add our custom combined columns 
    sheetHeaders = DataHeaders + [
        "Detail Footprint",
        "Detail State",
        "Detail Source IP",
        "Detail Source Port",
        "Detail Destination IP",
        "Detail Destination Port",
        "Sample Source IPs",
        "Sample Source Ports",
        "Sample Dest IPs",
        "Sample Dest Ports",
        "Sample Physical Ports",
        "Sample Vlan Tags",
        "Sample MPLS RD",
        "Sample Protocol"]
    #print(sheetHeaders)
    sheet.append(sheetHeaders)
    timer.lap("Headers Appended")

    #Make the headers bold
    for cell in sheet["1:1"]:
        cell.font = openpyxl.styles.Font(bold=True)
    timer.lap("Headers Bold")

     # Alignment and border shared by both styles
    top_align = openpyxl.styles.Alignment(vertical='top', wrapText=True)
    border_style = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(style='hair'),
        right=openpyxl.styles.borders.Side(style='hair'),
        top=openpyxl.styles.borders.Side(style='thin'),
        bottom=openpyxl.styles.borders.Side(style='thin')
    )
    # Fill for alternating rows
    alt_fill = openpyxl.styles.PatternFill(start_color=fillColor, end_color=fillColor, fill_type='solid')
    # Named styles
    style_name_base = "RowStyle_Normal"
    style_name_alt = "RowStyle_Alt"
    style_normal = openpyxl.styles.NamedStyle(name=style_name_base)
    style_normal.alignment = top_align
    style_normal.border = border_style
    style_alt = openpyxl.styles.NamedStyle(name=style_name_alt)
    style_alt.alignment = top_align
    style_alt.border = border_style
    style_alt.fill = alt_fill
    # Register styles once per workbook
    for style in [style_normal, style_alt]:
        if style.name not in wb.named_styles:
            wb.add_named_style(style)
    timer.lap("Styles defined")

    #Create an array of all entries in the input file.
    ##rawEntries = rawData.split("*********************************************************************\n")
    rawEntries = re.split(r'\n\*{69}(?:,*)\n',rawData)
    line_count = rawData.count('\n') + 1 if rawData else 0
    print(f"    {len(rawEntries)} entries found over {line_count} lines.")
    timer.total('ms')

    #Loop through each item of the array, processing the data within and creating a row in our output sheet.
    curRow = 2
    curLine=1
    format_set = False
    startTime = time.perf_counter()
    previous_entry = None
    for rawEntry in rawEntries:
        timer.reset()
        if len(rawEntry) > 0:
            #Populate entry with the desired data. See class clsEntry.
            entry = clsEntry(rawEntry, curLine)
            previous_entry = rawEntry
            timer.lap("Entry class created")

            if hasattr(entry,'error'):
                sheet.cell(row=curRow,column=1).value = entry.error
                curRow += 1
                continue
            
            #For troubleshooting, enable the following rows:
            #print(f'Processing Attack ID: {entry.defaultHeader[8]} S.No: {entry.defaultHeader}')
            #print(f'=============\n{rawEntry}\n==================')

            #Populate the cells in our spreadsheet row.
            for curColumn in range(len(DataHeaders)): #Populate columns A-Z
                if entry.data[curColumn]:
                    sheet.cell(row=curRow,column=curColumn + 1).value = entry.data[curColumn]
            sheet.cell(row=curRow,column=len(DataHeaders)+1).value = entry.footprint
            sheet.cell(row=curRow,column=len(DataHeaders)+2).value = entry.state
            sheet.cell(row=curRow,column=len(DataHeaders)+3).value = entry.sIP
            sheet.cell(row=curRow,column=len(DataHeaders)+4).value = entry.sPort
            sheet.cell(row=curRow,column=len(DataHeaders)+5).value = entry.dIP
            sheet.cell(row=curRow,column=len(DataHeaders)+6).value = entry.dPort
            sno = entry.data[0]
            sno = sno.replace('\n    ',',')
            timer.lap("First block written")
            if len(entry.samples) > 0:
                for sample in entry.samples: 
                    for i in range(8):
                        if len(sample) >= i + 1:
                            if not sheet.cell(row=curRow,column=len(DataHeaders)+7+i).value:
                                sheet.cell(row=curRow,column=len(DataHeaders)+7+i).value = sample[i]
                            else:
                                sheet.cell(row=curRow,column=len(DataHeaders)+7+i).value += "\n" + sample[i]
                        else:
                            sheet.cell(row=curRow,column=len(DataHeaders)+7+i).value = '\n    '.join(['Error',sheet.cell(row=curRow,column=33+i).value or '']).strip()
                #Remove duplicates from the samples we just added.
                for i in range(8):
                    curCell = sheet.cell(row=curRow,column=len(DataHeaders)+7+i)
                    curCell.value = '\n'.join(set(curCell.value.split('\n')))
            timer.lap("Samples parsed")
            #Replace commas with newlines 
            for i in [sheetHeaders.index("Detail Source IP")+1,
                      sheetHeaders.index("Detail Source Port")+1,
                      sheetHeaders.index("Detail Destination IP")+1,
                      sheetHeaders.index("Detail Destination Port")+1
                      ]:
                curCell = sheet.cell(row=curRow,column=i)
                curCell.value = '\n'.join(sorted(set(curCell.value.split(','))))
            timer.lap("Part 2")
            #Sort multiline entries. DataHeaders can't be multiline. The first two Detail headers aren't multiline. Apply to all headers that come 2 after the last dataheader.
            def custom_sort(item):
                if '.' in item: #IPv4
                    parts = item.split('.')
                    return [int(part) for part in parts]
                elif ':' in item: #IPv6
                    hextets = item.split(':')
                    if len(hextets) < 8:#Properly handle :: in ipv6 address
                        for index, value in enumerate(hextets):
                            if value == '':
                                hextets[index:index+1] = ['0'] * (9 - len(hextets))
                    
                    return [int(hextet,16) for hextet in hextets]
                elif item.isnumeric(): #Simple number
                    return [int(item)]
                else: #Text
                    return [1]
            #for i in range(len(DataHeaders)+4,len(sheetHeaders)+1):#range(29,41):#[29,30,31,32,33,34,35,36,37,38,39,40]:
            for i in [sheetHeaders.index("Sample Source IPs")+1,
                      sheetHeaders.index("Sample Source Ports")+1,
                      sheetHeaders.index("Sample Dest IPs")+1,
                      sheetHeaders.index("Sample Dest Ports")+1,
                      sheetHeaders.index("Sample Physical Ports")+1,
                      sheetHeaders.index("Sample Vlan Tags")+1,
                      sheetHeaders.index("Sample MPLS RD")+1,
                      sheetHeaders.index("Sample Protocol")+1
                      ]:
                curCell = sheet.cell(row=curRow,column=i)
                if curCell.value:
                    item = set(curCell.value.strip().split('\n'))
                    if len(item) > 1:
                        #print("---")
                        #print(curCell.value )
                        try:
                            curCell.value = '\n'.join(sorted(item, key=custom_sort))
                        except:
                            print(f"    Error sorting file: {file} s.no: {sno} row: {curRow} column {i}({openpyxl.utils.get_column_letter(i)})")
            timer.lap("Sorted")
            
            #Apply named style to the row
            style_to_use = style_name_alt if ColorAlternateRows and curRow % 2 == 0 else style_name_base
            for i in range(1, len(sheetHeaders) + 1):
                sheet.cell(row=curRow, column=i).style = style_to_use
            timer.lap("Formatting Finished")

            #Format date columns as dates
            for i in [sheetHeaders.index("Start Time")+1,sheetHeaders.index("End Time")+1]:
                cell = sheet.cell(row=curRow,column=i)
                
                try:
                    #Columns B and C are dates. Lets convert it to a proper date format.
                    if cell.value:
                        # Parse the date time string to a datetime object
                        #datetime_obj = datetime.strptime(cell.value.strip(), dp_one_line_date_format)
                        #cell.value = datetime_obj.strftime("%Y-%m-%d %H:%M:%S")
                        lines = str(cell.value).strip().splitlines()
                        new_lines = []
                        for line in lines:
                            dt = datetime.strptime(line.strip(), dp_one_line_date_format)
                            new_lines.append(dt.strftime("%Y-%m-%d %H:%M:%S"))
                        cell.value = "\n".join(new_lines)
                        # Convert the datetime object to the desired format
                        cell.number_format = "YYYY-MM-DD HH:MM:SS"
                        
                except Exception as err:
                    print(f"  Error processing date at file: {file} s.no: {sno} row: {curRow} column: {i} expected: {dp_one_line_date_format} actual: '{cell.value}'")
                    print(f"    Details: {str(err).replace(chr(10), ' | ')}")
                    raise ValueError("Err5:Bad date")
            timer.lap("Dates formatted")
            timer.total("ms")
        #Output current progress once per hundred entried processed
        if (curRow - 1) % 100 == 0:
            endTime = time.perf_counter()
            print(f"    Processed {curRow - 1} of {len(rawEntries)} entries in {(endTime - startTime) * 1000:.2f} ms", end='\r', flush=True)
            startTime = time.perf_counter()
        curRow += 1
        curLine += len(rawEntry.split('\n')) + 1
    print(f"    Processed {curRow-2} of {len(rawEntries)} entries")
    # Auto-fit columns to fit the data
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        cellsWithData=0
        for cell in column:
            try:
                #wrapped_text = textwrap.wrap(str(cell.value), width=60)  # Adjust the width as needed
                wrapped_text = str(cell.value).strip()
                if cell.value and len(cell.value) > 0 and cell.value != 'N/A':
                    cellsWithData += 1
                for line in wrapped_text.splitlines():
                    if len(line) > max_length:
                        max_length = len(line)
                    
            except Exception as e:
                print(e)
                pass

            #Format numeric cells as numbers.
            if type(cell.value) is str:
                if cell.value.replace(".", "", 1).isnumeric():
                    cell.value = float(cell.value)
        
        

        if hideEmptyColumns == False or cellsWithData > 1:
            adjusted_width = (max_length + 2) * 1.01  # Adjust the multiplier as needed
            if adjusted_width > maxWidth: #maxWidth set near the top of the script.
                adjusted_width = maxWidth
            sheet.column_dimensions[column_letter].width = adjusted_width
        else: 
            #All cells in the column are empty. Hide the column.
            sheet.column_dimensions[column_letter].hidden = True
            
    #Freeze the header row
    sheet.freeze_panes = sheet['A2']

    #Save the worksheet
    retry=True
    while (retry == True):
        try:
            outFile = file.replace(".csv",".xlsx")
            print("    Saving to " + output_path + outFile)
            wb.save(output_path + outFile)
            print("    Saved successfully!")
            retry = False
        except Exception as e:
            print(f'\n  Error writing to {output_path + outFile}\n    {e}')
            print("  Please make sure the document is not currently open!")
            print("  Press enter to retry. Press any other key to abort")
            strInput = input()
            if len(strInput) > 0:
                retry = False
        except:
            print("  Write failed")
            retry = False
    

if not os.path.exists(input_path):
    print("input subfolder not found. It will be created for you.")
    os.makedirs(input_path)

if not os.path.exists(output_path):
    os.makedirs(output_path)

for path, dir, files in os.walk(input_path):
    if path in ['./input/noprocess','./input/ignore','./input/old']:
        continue
    if len(files) == 0:
        print("Please place DefenseProForensicReport.csv files in the ./input/ folder and rerun the script.")
    for file in files:
        if file.endswith(".tgz") or file.endswith(".zip"):
            try:
                print("zip/tgz file support to be added later. let Steve Harris know if this is a feature that would be helpful for you.\r    " + file)
            except Exception as err:
                print(f'Error processing {input_path + file} {err}')
        elif file.endswith(".csv"):
            print("Processing" + input_path + file)
            with open(input_path + file, 'r') as f:
                rawData = f.read()
                outFile = file.replace(".csv",".xlsx")
                if replaceExistingFile or not os.path.exists(output_path + outFile):
                    processData(rawData)
                else:
                    print(f"  Output file: {output_path + outFile} already exists. Skipping the processing of {input_path + file}")    
